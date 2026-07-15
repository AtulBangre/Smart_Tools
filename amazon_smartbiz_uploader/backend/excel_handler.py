import openpyxl
from openpyxl.utils import get_column_letter

def generate_smartbiz_excel(scraped_data: list, template_path: str, output_path: str) -> bool:
    """
    Loads the template, appends scraped_data to 'bulk_upload_template', and saves to output_path.
    """
    try:
        wb = openpyxl.load_workbook(template_path)
        
        # Access the specific sheet
        if 'bulk_upload_template' not in wb.sheetnames:
            print("Error: Sheet 'bulk_upload_template' not found in template.")
            return False
            
        ws = wb['bulk_upload_template']
        
        # Find the first empty row in the sheet
        # Usually row 1 is headers (or row 1-3). We start checking from row 2
        start_row = 1
        while ws.cell(row=start_row, column=4).value or ws.cell(row=start_row, column=1).value or ws.cell(row=start_row, column=2).value: 
            # We assume if SKU ID or Variant ID or Product Name is present, it's not empty
            start_row += 1
            
        # If it's a completely fresh template, headers might be on row 1
        # Let's ensure start_row is at least 2
        if start_row == 1:
            start_row = 2
            
        current_row = start_row
        
        for data in scraped_data:
            # Column mapping (1-indexed in openpyxl)
            # 1: SKU ID
            # 2: Variant ID
            # 3: Custom SKU (User input) -> C
            # 4: Product Name (Scraped) -> D
            # 5: MRP (Scraped) -> E
            # 6: Selling Price (Scraped) -> F
            # 7: Business Category (User input) -> G
            # 8: Product Category (User input) -> H
            # 9: Product Description (Scraped) -> I
            
            ws.cell(row=current_row, column=3).value = data.get('custom_sku', '')
            ws.cell(row=current_row, column=4).value = data.get('name', '')
            ws.cell(row=current_row, column=5).value = data.get('mrp', '')
            ws.cell(row=current_row, column=6).value = data.get('selling_price', '')
            ws.cell(row=current_row, column=7).value = data.get('business_category', '')
            ws.cell(row=current_row, column=8).value = data.get('product_category', '')
            ws.cell(row=current_row, column=9).value = data.get('description', '')
            ws.cell(row=current_row, column=10).value = data.get('variant_relationship', '')
            ws.cell(row=current_row, column=11).value = data.get('size', '')
            ws.cell(row=current_row, column=13).value = data.get('color_name', '')
            ws.cell(row=current_row, column=14).value = data.get('best_seller', 'No')
            
            images = data.get('images', [])
            for idx, img_url in enumerate(images):
                if idx < 6:
                    # Images start at column 16 (Product Image1) up to 21 (Product Image6)
                    ws.cell(row=current_row, column=16 + idx).value = img_url
            
            ws.cell(row=current_row, column=23).value = data.get('seo_title', '')
            ws.cell(row=current_row, column=24).value = data.get('seo_description', '')
            
            current_row += 1
            
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"Error handling Excel file: {e}")
        return False
