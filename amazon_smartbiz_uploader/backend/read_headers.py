import openpyxl
wb = openpyxl.load_workbook("Smartbiz_Upload_Generated.xlsx")
ws = wb['bulk_upload_template']
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    headers.append(f"{col}: {val}")
print("\n".join(headers))
